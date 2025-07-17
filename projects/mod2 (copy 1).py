import os
import time
import sys
# from tkinter import *
# from tkinter import ttk
import configparser as cp
from barcode import EAN8
from barcode.writer import ImageWriter 
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import csv
import shutil
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtWidgets import QApplication, QMainWindow, QPushButton, QLineEdit, QLabel, QVBoxLayout, QWidget, QComboBox

#Путь до исходной таблицы
pathToTable = ""
#Назвение файла в прошлом каталоге
nameOfTable = "table.xlsx"
#Путь до конфига
pathToConfig = "config.ini"
#Путь до сохранения файлов бланков
pathToSave = "pictures"
#Пути до шаблонов
nameOfTitle = "list.png"
nameOfAns = "Ans.png"

Students = []
Subjects = set()

class Student:

    def __init__(self, Code, FirstName, LastName, SurName, Profile, Class, Subject):
        self.Code = Code
        self.FirstName = FirstName
        self.LastName = LastName
        self.SurName = SurName
        self.Profile = Profile
        self.Class = Class
        self.Subject = Subject

def reopen(imgX, code):
    imgX.save(f"{pathToSave}{code}")
    imgX1 = Image.open(f"{pathToSave}{code}.png")
    return imgX1

def xlsx_2_csv(name, out_name):
    wb = load_workbook(filename=name)
    ws = wb.active
    csv_data = [list(value) for value in ws.iter_rows(values_only=True)]
    with open(out_name, "w") as file:
        writer = csv.writer(file, delimiter=",")
        for line in csv_data:
            writer.writerow(line)

def init():
    global Students, Subjects
    Students = []
    
    #shutil.copy(f"{nameOfTable}", "table1.xlsx")
    #xlsx_2_csv("table1.xlsx", "table.csv")

    config = cp.ConfigParser()
    config.read(pathToConfig)
    col = config["Coloumns"]
    fn, ln, sn, p, c, s = map(int, [col["FirstName"], col["LastName"], col["SurName"], col["Profile"], col["Class"], col["Subject"]])
    nums = config["Nums"]


    numOfPerson = dict()
    with open(pathToTable + nameOfTable, mode = "r", encoding = "utf-8", newline = "") as r_file:
        file_reader = list(csv.reader(r_file))
        # print(file_reader)
        for a in file_reader[int(nums["StartStr"]):]:
        # for i in range(int(nums["StartStr"]), int(nums["NumPeople"])):
            # a = list(map(str, f[i].split(','))) 
            if numOfPerson.get(a[c]) == None: numOfPerson[a[c]] = 1
            else: numOfPerson[a[c]] += 1

            code = str(a[c][:-1]).zfill(4) + str(numOfPerson[a[c]]).zfill(3)
            print(code, a[s])
            Students.append(Student(code, a[fn], a[ln], a[sn], a[p], a[c], a[s].lower()))
            Subjects.add(a[s].lower())

#Генерация бланков
def generateList(Subject):
    global Students
    if not (os.path.exists(nameOfTitle) or os.path.exists(nameOfAns)):
        errw = ErrorWindow()
        errw.show()
        return

    for Student in Students:
        if Student.Subject != Subject: continue
        # print(Student.Subject, Subject)
        Title = Image.open(nameOfTitle)
        Ans = Image.open(nameOfAns)
        
        bar = EAN8(f"{Student.Code}", writer = ImageWriter())
        bar = reopen(bar, Student.Code)
        bar = bar.crop((0, 0, 350, 200))
        Title.paste(bar, (0, 100))
        Ans.paste(bar, (0, 100))
        
        font = ImageFont.truetype('calibri.ttf', size=100)
        drawText = ImageDraw.Draw(Title);
        drawText.text((1000, 850),
                      f"{Subject[0].upper() + Subject[1:]}\n\n{Student.FirstName}\n\n{Student.LastName}\n\n{Student.SurName}\n\n{Student.Class}",
                      font = font,
                      fill=("#1C0606"))
        
        Title.save(f"{pathToSave}{Student.Code}list.png")
        Ans.save(f"{pathToSave}{Student.Code}ans.png")

#Генерация таблицы резов
def generateTable(Subject):
    global Students
    with open("codes.csv", mode="w", encoding='utf-8') as w_file:
        file_writer = csv.writer(w_file, delimiter = ",", lineterminator = "\r")
        file_writer.writerow(["Предмет", "Код", "Баллы"])
        for Student in Students:
            if(Student.Subject[:len(Subject)] != Subject): continue
            file_writer.writerow([Student.Subject, Student.Code])

def generateTableWithNames(Subject):
    global Students
    with open("codes.csv", mode="r+", encoding='utf-8') as r_file:
        with open("results.csv", mode = "w", encoding='utf-8') as w_file:
            fr = list(csv.reader(r_file))
            fw = csv.writer(w_file, delimiter = ",", lineterminator = "\r")
            fw.writerow(["Предмет", "ФИО", "Баллы"])
            i = 1
            for Student in Students:
                if(Student.Subject[:len(Subject)] != Subject): continue
                fr[i][1] = Student.LastName + ' ' + Student.FirstName
                fw.writerow(fr[i])

class MainWindow(QMainWindow):

    def __init__(self, parent = None):
        super(MainWindow, self).__init__(parent)
        self.selectedSubject = ""
        layout = QVBoxLayout()
        self.setWindowTitle("Generator")
        self.setMinimumSize(QSize(600, 300))
        
        self.text1 = QLabel("Путь до таблицы")
        font = self.text1.font()
        font.setPointSize(10)
        self.text1.setFont(font)
        self.text1.setAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)

        self.widgetPath = QLineEdit()
        self.widgetPath.setPlaceholderText("Если нужно сохранить в папке с приложением, оставьте пустым")

        self.widgetPath = QLineEdit()
        self.widgetPath.setPlaceholderText("Если файл лежит в папке с приложением, оставьте пустым")
        # self.widgetPath.returnPressed.connect(self.return_pressed_line_1)

        self.text2 = QLabel("Имя таблицы")
        self.text2.setFont(font)
        self.text2.setAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)

        self.widgetName = QLineEdit()
        # self.widgetName.returnPressed.connect(self.return_pressed_line_2)

        self.initial = QPushButton("Импортировать из таблицы")
        self.initial.setCheckable(True)
        self.initial.clicked.connect(self.Initial)

        self.text3 = QLabel("Путь для coхранения бланков")
        self.text3.setFont(font)
        self.text3.setAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)

        self.blankPath = QLineEdit()
        self.blankPath.setPlaceholderText("Если бланки нужно сохранить в папку с приложением, оставьте пустым")
        
        self.gen = QPushButton("Сгенерировать бланки")
        self.gen.setCheckable(True)
        self.gen.clicked.connect(self.OnGenerate)

        self.subjects = QComboBox()
        global Subjects
        self.subjects.currentTextChanged.connect(self.SelectSub)
        # self.subjects.setMaxCount(len(Subjects))

        self.genTable = QPushButton("Сгенерировать таблицу кодов")
        self.genTable.setCheckable(True)
        self.genTable.clicked.connect(self.OnGenTable)

        self.genTNames = QPushButton("Заменить коды на имена")
        self.genTNames.setCheckable(True)
        self.genTNames.clicked.connect(self.ChangeCodes)

        """widget.selectionChanged.connect(self.selection_changed)
        widget.textChanged.connect(self.text_changed)
        widget.textEdited.connect(self.text_edited)"""

        layout.addWidget(self.text1)
        layout.addWidget(self.widgetPath)
        layout.addWidget(self.text2)
        layout.addWidget(self.widgetName)
        layout.addWidget(self.initial)
        layout.addWidget(self.subjects)
        layout.addWidget(self.text3)
        layout.addWidget(self.blankPath)
        layout.addWidget(self.gen)
        layout.addWidget(self.genTable)
        layout.addWidget(self.genTNames)

        container = QWidget()
        container.setLayout(layout)
        
        self.setCentralWidget(container)

    def Initial(self):
        global pathToTable, nameOfTable, Subjects
        pathToTable = self.widgetPath.text()
        nameOfTable = self.widgetName.text()
        if not pathToTable: pathToTable = "./"
        if pathToTable[-1] != "/": pathToTable += "/"
        if not (os.path.exists(pathToTable + nameOfTable)):
            errw = ErrorWindow(self)
            errw.show()
            print(1)
        else:
            init()
            self.subjects.clear()
            self.subjects.addItems(Subjects)

    def OnGenerate(self):
        global pathToSave
        pathToSave = self.blankPath.text()
        if not pathToSave: pathToSave = "./"
        if pathToSave[-1] != "/": pathToSave += "/"
        if not (os.path.exists(nameOfTitle) or os.path.exists(nameOfAns)):
            errw = ErrorWindow()
            errw.show()
            return
        else:
            generateList(self.selectedSubject)

    def SelectSub(self, s):
        self.selectedSubject = s
        print(self.selectedSubject)

    def OnGenTable(self):
        generateTable(self.selectedSubject)

    def ChangeCodes(self):
        generateTableWithNames(self.selectedSubject)

class ErrorWindow(QMainWindow):

    def __init__(self, parent = None):
        super(ErrorWindow, self).__init__(parent)
        self.text = QLabel("Файл не найден или перемещен")
        self.setCentralWidget(self.text)
         


def main():
    app = QApplication([])

    window = MainWindow()
    window.show()

    sys.exit(app.exec())

if __name__ == "__main__":
    main()
